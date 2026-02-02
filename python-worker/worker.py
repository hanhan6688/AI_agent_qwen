import redis
import requests
import json
import time
import logging
import os
import sys
import re
import threading
import traceback
from datetime import datetime
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv

# 加载环境变量
load_dotenv()

# 配置日志
log_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logs')
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, 'worker.log')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Redis配置
REDIS_HOST = os.getenv('REDIS_HOST', 'localhost')
REDIS_PORT = int(os.getenv('REDIS_PORT', 6379))
REDIS_PASSWORD = os.getenv('REDIS_PASSWORD', '')
MINERU_API_KEY = os.getenv('MINERU_API_KEY', '')
DASHSCOPE_API_KEY = os.getenv('DASHSCOPE_API_KEY', '')
QWEN_API_KEY = os.getenv('QWEN_API_KEY', '')
BACKEND_API_URL = os.getenv('BACKEND_API_URL', 'http://localhost:8080')

logger.info(f"Redis配置: host={REDIS_HOST}, port={REDIS_PORT}, password={'[已设置]' if REDIS_PASSWORD else '[未设置]'}")

# 队列名称
PENDING_QUEUE = "task:pending"
PROCESSING_QUEUE = "task:processing"

# Qwen-VL 配置
MODEL_NAME = "qwen-vl-max"
MAX_CONTEXT_LENGTH = 150000
MAX_IMAGES = 15
MAX_TPM = 1000000
TOKEN_BUCKET = MAX_TPM
LAST_REFILL_TIME = time.time()
TOKEN_LOCK = threading.Lock()
POLL_INTERVAL = 10
BATCH_SIZE = 200


class RedisQueue:
    def __init__(self, host, port, password=''):
        # 如果密码为空字符串，则不传递password参数（无密码连接）
        if password and password.strip():
            self.redis = redis.Redis(
                host=host,
                port=port,
                password=password,
                decode_responses=True
            )
        else:
            self.redis = redis.Redis(
                host=host,
                port=port,
                decode_responses=True
            )
        # 测试连接
        try:
            self.redis.ping()
            logger.info("Redis连接成功")
        except Exception as e:
            logger.error(f"Redis连接失败: {e}")
            raise

    def get_next_task(self) -> Optional[str]:
        """从待处理队列获取下一个任务"""
        task_id = self.redis.lpop(PENDING_QUEUE)
        if task_id:
            # 将任务添加到处理中队列
            self.redis.hset(PROCESSING_QUEUE, task_id, str(time.time()))
            logger.info(f"从队列获取任务: {task_id}")
        return task_id

    def mark_completed(self, task_id: str):
        """标记任务为完成"""
        self.redis.hdel(PROCESSING_QUEUE, task_id)
        logger.info(f"任务 {task_id} 标记为完成")

    def mark_failed(self, task_id: str):
        """标记任务为失败"""
        self.redis.hdel(PROCESSING_QUEUE, task_id)
        logger.error(f"任务 {task_id} 标记为失败")


class BackendClient:
    def __init__(self, base_url: str):
        self.base_url = base_url

    def get_task(self, task_id: str) -> Optional[Dict]:
        """获取任务详情"""
        try:
            url = f"{self.base_url}/api/tasks/{task_id}"
            logger.info(f"获取任务详情: {url}")
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                return data.get('data')
            else:
                logger.error(f"获取任务失败: {response.status_code} - {response.text}")
                return None
        except Exception as e:
            logger.error(f"获取任务 {task_id} 失败: {e}")
            return None

    def update_task_status(self, task_id: str, status: str) -> bool:
        """更新任务状态 - 调用Spring Boot API"""
        try:
            url = f"{self.base_url}/api/tasks/{task_id}/status"
            payload = {"status": status}
            logger.info(f"更新任务状态: {url} -> {status}")
            response = requests.put(url, json=payload, timeout=10)
            if response.status_code == 200:
                logger.info(f"任务 {task_id} 状态更新成功: {status}")
                return True
            else:
                logger.error(f"更新任务状态失败: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            logger.error(f"更新任务状态失败: {e}")
            return False

    def update_task_excel(self, task_id: str, excel_path: str) -> bool:
        """更新任务Excel路径 - 调用Spring Boot API"""
        try:
            url = f"{self.base_url}/api/tasks/{task_id}/excel"
            payload = {"excelPath": excel_path}
            logger.info(f"更新任务Excel路径: {url}")
            response = requests.put(url, json=payload, timeout=10)
            if response.status_code == 200:
                logger.info(f"任务 {task_id} Excel路径更新成功")
                return True
            else:
                logger.error(f"更新任务Excel路径失败: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            logger.error(f"更新任务Excel路径失败: {e}")
            return False

    def update_task_error(self, task_id: str, error_message: str) -> bool:
        """更新任务错误信息 - 调用Spring Boot API"""
        try:
            url = f"{self.base_url}/api/tasks/{task_id}/error"
            payload = {"errorMessage": error_message}
            logger.error(f"更新任务错误信息: {url} - {error_message}")
            response = requests.put(url, json=payload, timeout=10)
            if response.status_code == 200:
                logger.info(f"任务 {task_id} 错误信息更新成功")
                return True
            else:
                logger.error(f"更新任务错误信息失败: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            logger.error(f"更新任务错误信息失败: {e}")
            return False


class PDFProcessor:
    """PDF转Markdown处理器，使用MinERU API"""
    
    def __init__(self):
        self.mineru_api_url = "https://mineru.net/api/v4"
        self.token = MINERU_API_KEY
        self.headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.token}"
        }

    def pdf_to_markdown(self, file_path: str, output_dir: str) -> str:
        """
        将PDF文件转换为Markdown
        返回Markdown文件路径
        """
        try:
            import zipfile
            
            pdf_path = Path(file_path)
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF文件不存在: {file_path}")

            logger.info(f"开始处理PDF: {file_path}")

            # 1. 获取上传URL
            file_data = {
                "name": pdf_path.name,
                "is_ocr": True,
                "data_id": f"{pdf_path.stem}-pdf-id"
            }

            payload = {
                "enable_formula": True,
                "language": "ch",
                "enable_table": True,
                "files": [file_data]
            }

            url = f"{self.mineru_api_url}/file-urls/batch"
            response = requests.post(url, headers=self.headers, json=payload, timeout=30)

            if response.status_code != 200:
                raise RuntimeError(f"获取上传URL失败: HTTP {response.status_code}")

            resp = response.json()
            if resp.get("code") != 0:
                error_msg = resp.get("msg", "未知错误")
                raise RuntimeError(f"获取上传URL失败: {error_msg}")

            # 2. 上传PDF
            batch_id = resp["data"]["batch_id"]
            upload_url = resp["data"]["file_urls"][0]

            logger.info(f"上传PDF到MinERU, batch_id: {batch_id}")

            with open(pdf_path, "rb") as f:
                upload_response = requests.put(upload_url, data=f, timeout=300)
                if upload_response.status_code != 200:
                    raise RuntimeError(f"PDF上传失败: HTTP {upload_response.status_code}")

            # 3. 轮询等待处理完成
            result_url = f"{self.mineru_api_url}/extract-results/batch/{batch_id}"
            max_wait = 600  # 最多等待10分钟
            wait_count = 0

            while wait_count < max_wait:
                time.sleep(POLL_INTERVAL)
                wait_count += POLL_INTERVAL

                result_response = requests.get(result_url, headers=self.headers, timeout=30)
                if result_response.status_code != 200:
                    continue

                result_data = result_response.json()
                if result_data.get("code") != 0:
                    continue

                results = result_data["data"]["extract_result"]
                if not results:
                    continue

                state = results[0].get("state")
                logger.info(f"处理状态: {state}")

                if state == "done":
                    # 4. 下载并解压结果
                    zip_url = results[0].get("full_zip_url", "").strip()
                    if not zip_url:
                        raise RuntimeError("未获取到结果下载URL")

                    # 创建输出目录
                    output_path = Path(output_dir)
                    output_path.mkdir(parents=True, exist_ok=True)

                    # 下载ZIP
                    zip_path = output_path / f"{pdf_path.stem}.zip"
                    logger.info(f"下载结果到: {zip_path}")

                    download_response = requests.get(zip_url, stream=True, timeout=300)
                    download_response.raise_for_status()

                    with open(zip_path, "wb") as f:
                        for chunk in download_response.iter_content(chunk_size=8192):
                            f.write(chunk)

                    # 解压
                    extract_dir = output_path / pdf_path.stem
                    extract_dir.mkdir(exist_ok=True)
                    with zipfile.ZipFile(zip_path) as zf:
                        zf.extractall(extract_dir)

                    # 删除ZIP文件
                    zip_path.unlink()

                    # 查找markdown文件
                    md_files = list(extract_dir.glob("**/*.md"))
                    if not md_files:
                        raise RuntimeError("未找到生成的Markdown文件")

                    md_path = md_files[0]
                    logger.info(f"PDF转换完成: {md_path}")

                    return str(md_path)

                elif state in ("failed", "error"):
                    error_msg = results[0].get("err_msg", "未知错误")
                    raise RuntimeError(f"PDF处理失败: {error_msg}")

            raise RuntimeError("处理超时")

        except Exception as e:
            logger.error(f"PDF转Markdown失败: {e}")
            raise


def refill_token_bucket():
    """补充令牌桶"""
    global TOKEN_BUCKET, LAST_REFILL_TIME
    current_time = time.time()
    elapsed = current_time - LAST_REFILL_TIME
    if elapsed > 60:
        TOKEN_BUCKET = MAX_TPM
        LAST_REFILL_TIME = current_time
    else:
        refill_amount = (elapsed / 60) * MAX_TPM
        TOKEN_BUCKET = min(MAX_TPM, TOKEN_BUCKET + refill_amount)
        LAST_REFILL_TIME = current_time

def wait_for_tokens(required_tokens):
    """等待直到令牌桶中有足够令牌"""
    global TOKEN_BUCKET
    with TOKEN_LOCK:
        refill_token_bucket()
        while TOKEN_BUCKET < required_tokens:
            deficit = required_tokens - TOKEN_BUCKET
            wait_seconds = (deficit / MAX_TPM) * 60 + 0.1
            logger.info(f"TPM不足，需要等待{wait_seconds:.2f}秒 (需求:{required_tokens} 可用:{TOKEN_BUCKET:.0f})")
            time.sleep(wait_seconds)
            refill_token_bucket()

        TOKEN_BUCKET -= required_tokens
        logger.info(f"扣除{required_tokens} tokens，剩余:{TOKEN_BUCKET:.0f}")


class DataExtractor:
    """从Markdown文档中提取数据的处理器"""
    
    def __init__(self):
        try:
            from dashscope import MultiModalConversation
            import dashscope
            dashscope.api_key = DASHSCOPE_API_KEY
            self.MultiModalConversation = MultiModalConversation
            logger.info("Qwen API初始化成功")
        except ImportError as e:
            logger.error(f"dashscope模块未安装: {e}")
            self.MultiModalConversation = None

        # 加载提示词
        prompt_path = os.path.join(os.path.dirname(__file__), '..', 'prompt.txt')
        try:
            self.prompt_txt = open(prompt_path, encoding="utf-8").read()
        except FileNotFoundError:
            logger.warning("未找到prompt.txt文件，使用默认提示词")
            self.prompt_txt = "请从文档中提取结构化数据"

    def preprocess_context(self, context: str) -> str:
        """文本预处理 - 移除不需要的部分"""
        sections_to_remove = [
            r'references?',
            r'acknowledg?e?ments?',
            r'data availability',
            r'declaration of competing interest',
            r'conflict of interest',
            r'funding',
            r'appendix',
        ]
        pattern = r'(?i)\n#*\s*(' + '|'.join(sections_to_remove) + r')[\s\S]*?(\n#|$)'
        context = re.sub(pattern, '', context)
        context = re.sub(r'(?i)(\n|^)\s*acknowledg?e?ments?[\s\S]*?(\n#|$)', '', context)
        return context[:MAX_CONTEXT_LENGTH]

    def try_repair_json(self, json_str: str):
        """针对Qwen-VL输出的JSON修复"""
        cleaned = json_str.strip()

        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()

        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            logger.warning(f"JSON解析失败，尝试修复")
            cleaned = re.sub(r',\s*([}\]])', r'\1', cleaned)
            cleaned = re.sub(r'([{\[])\s*,', r'\1', cleaned)

            try:
                return json.loads(cleaned)
            except json.JSONDecodeError:
                return None

    def build_messages(self, text: str, img_abs_paths: List[str]) -> List[Dict]:
        """为Qwen-VL构建消息格式"""
        system_msg = {
            "role": "system",
            "content": [{"text": self.prompt_txt}]
        }

        user_content = [{"text": text}]
        for img_path in img_abs_paths:
            user_content.append({"image": f"file://{img_path}"})

        user_msg = {
            "role": "user",
            "content": user_content
        }

        return [system_msg, user_msg]

    def extract_once(self, md_file: str) -> Tuple[str, Any]:
        """使用Qwen-VL进行提取"""
        if not self.MultiModalConversation:
            return ("error", "Qwen API未初始化")

        try:
            # 1. 读取并预处理文本
            text = open(md_file, encoding="utf-8").read()
            text = self.preprocess_context(text)

            # 2. 解析图片路径
            md_dir = Path(md_file).parent
            fig_imgs = []

            pattern = re.compile(
                r'!\[\]\(images/([^)]+)\)[\s\S]*?(Fig\.|Figure|图)\s?\d+[\.\d]*\..*?(\n|$)',
                re.IGNORECASE | re.MULTILINE
            )

            for m in pattern.finditer(text):
                img_name = m.group(1)
                rel_path = f"images/{img_name}"
                abs_path = (md_dir / rel_path).resolve()
                if abs_path.exists():
                    fig_imgs.append(str(abs_path))
                    logger.info(f"找到带Fig.描述的图片: {img_name}")

            if not fig_imgs:
                img_dir = md_dir / "images"
                if img_dir.exists():
                    for p in img_dir.glob("*"):
                        if p.suffix.lower() in {".jpg", ".jpeg", ".png", ".bmp", ".gif"} and "fig" in p.stem.lower():
                            fig_imgs.append(str(p.resolve()))

            abs_imgs = fig_imgs[:MAX_IMAGES]
            logger.info(f"使用 {len(abs_imgs)} 张图片")

            # 3. 估算token并调用API
            estimated_tokens = len(text) // 3.5 + len(abs_imgs) * 1000

            max_retries = 3
            rsp = None
            for attempt in range(max_retries):
                try:
                    wait_for_tokens(estimated_tokens)
                    rsp = self.MultiModalConversation.call(
                        model=MODEL_NAME,
                        messages=self.build_messages(text, abs_imgs),
                        temperature=0.1,
                        max_tokens=4000
                    )
                    if rsp.status_code == 200:
                        break
                    else:
                        logger.warning(f"API返回状态码异常: {rsp.status_code} (尝试 {attempt+1}/{max_retries})")
                except Exception as e:
                    logger.warning(f"API调用失败 (尝试 {attempt+1}/{max_retries}): {str(e)}")
                    if attempt < max_retries - 1:
                        wait_time = 2 ** attempt
                        logger.info(f"等待 {wait_time} 秒后重试...")
                        time.sleep(wait_time)

            if not rsp or rsp.status_code != 200:
                error_msg = f"API错误: {getattr(rsp, 'message', 'Unknown error')}" if rsp else "API调用失败"
                raise RuntimeError(error_msg)

            # 4. 解析返回结果
            content = rsp.output.choices[0].message.content
            if isinstance(content, list) and content and "text" in content[0]:
                json_str = content[0]["text"]

                try:
                    return ("success", json.loads(json_str))
                except json.JSONDecodeError:
                    pass

                repaired_obj = self.try_repair_json(json_str)
                if repaired_obj is not None:
                    return ("success", repaired_obj)

                return ("partial_data", json_str)
            else:
                return ("error", "API返回格式错误")

        except Exception as e:
            logger.error(f"处理失败: {e}")
            return ("error", str(e))

    def extract_from_markdown(self, markdown_path: str, extract_fields: str = '{}') -> str:
        """从Markdown文件提取数据，返回JSON字符串"""
        try:
            status, result = self.extract_once(markdown_path)

            if status == "success":
                return json.dumps(result, ensure_ascii=False, indent=2)
            else:
                raise Exception(f"提取失败: {result}")
        except Exception as e:
            logger.error(f"从Markdown提取数据失败: {e}")
            raise



class TaskWorker:
    """任务处理Worker，从Redis队列获取任务并处理"""
    
    def __init__(self):
        self.queue = RedisQueue(REDIS_HOST, REDIS_PORT, REDIS_PASSWORD)
        self.backend = BackendClient(BACKEND_API_URL)
        self.pdf_processor = PDFProcessor()
        self.data_extractor = DataExtractor()

        # 设置本地数据目录
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.processed_dir = os.path.join(project_root, 'data', 'processed')
        self.excel_dir = os.path.join(project_root, 'data', 'excel')

        # 确保目录存在
        os.makedirs(self.processed_dir, exist_ok=True)
        os.makedirs(self.excel_dir, exist_ok=True)

        logger.info("TaskWorker初始化完成")

    def process_task(self, task_id: str):
        """处理单个任务"""
        logger.info(f"开始处理任务: {task_id}")
        start_time = time.time()

        try:
            # 1. 更新任务状态为处理中
            self.backend.update_task_status(task_id, "PROCESSING")

            # 2. 获取任务详情
            task = self.backend.get_task(task_id)
            if not task:
                raise Exception("任务不存在")

            file_path = task.get('filePath')
            task_name = task.get('taskName', f'task_{task_id}')
            extract_fields = task.get('extractFields', '{}')

            if not file_path or not os.path.exists(file_path):
                raise Exception(f"文件不存在: {file_path}")

            logger.info(f"任务信息 - ID: {task_id}, 文件: {file_path}")

            # 3. PDF转Markdown
            logger.info("步骤1: PDF转Markdown")
            markdown_path = self.pdf_processor.pdf_to_markdown(file_path, self.processed_dir)

            # 4. 从Markdown提取数据
            logger.info("步骤2: 从Markdown提取数据")
            json_result = self.data_extractor.extract_from_markdown(markdown_path, extract_fields)

            # 5. 生成Excel
            logger.info("步骤3: 生成Excel文件")
            excel_path = self._generate_excel(json_result, task_name)

            # 6. 更新任务状态为完成（包括Excel路径）
            self.backend.update_task_excel(task_id, excel_path)
            # 从Redis处理中队列移除任务
            self.queue.mark_completed(task_id)

            elapsed = time.time() - start_time
            logger.info(f"任务 {task_id} 处理完成，耗时: {elapsed:.2f}秒，Excel: {excel_path}")

        except Exception as e:
            logger.error(f"任务 {task_id} 处理失败: {e}")
            logger.error(traceback.format_exc())
            self.backend.update_task_error(task_id, str(e))
            # 从Redis处理中队列移除任务
            self.queue.mark_failed(task_id)

    def _generate_excel(self, json_result: str, task_name: str) -> str:
        """生成Excel文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

            # 解析JSON
            data = json.loads(json_result)
            if isinstance(data, dict):
                data = [data]

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "提取结果"

            # 设置表头样式
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 写入表头
            headers = list(data[0].keys()) if data else []
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = header_alignment

            # 写入数据
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, (key, value) in enumerate(row_data.items(), start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if value is not None else ""
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 保存文件
            excel_filename = f"{task_name}_{int(time.time())}.xlsx"
            excel_path = os.path.join(self.excel_dir, excel_filename)
            wb.save(excel_path)

            logger.info(f"Excel文件生成完成: {excel_path}")
            return excel_path

        except Exception as e:
            logger.error(f"生成Excel失败: {e}")
            raise

    def run(self):
        """运行Worker主循环"""
        logger.info("=" * 50)
        logger.info("Worker启动，等待任务...")
        logger.info("=" * 50)

        while True:
            try:
                # 从队列获取任务
                task_id = self.queue.get_next_task()

                if task_id:
                    # 处理任务
                    self.process_task(task_id)
                else:
                    # 没有任务，休眠1秒
                    time.sleep(1)

            except KeyboardInterrupt:
                logger.info("收到停止信号，Worker退出")
                break
            except Exception as e:
                logger.error(f"Worker运行出错: {e}")
                logger.error(traceback.format_exc())
                time.sleep(5)


if __name__ == "__main__":
    try:
        worker = TaskWorker()
        worker.run()
    except KeyboardInterrupt:
        logger.info("Worker停止")
    except Exception as e:
        logger.error(f"Worker异常退出: {e}")
        logger.error(traceback.format_exc())
        sys.exit(1)